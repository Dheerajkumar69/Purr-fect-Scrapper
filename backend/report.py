"""
report.py — Generate HTML and JSON reports from merged scraping results.

Outputs:
  /output_dir/reports/{job_id}.json  — full machine-readable merged output
  /output_dir/reports/{job_id}.html  — human-readable HTML report
"""

from __future__ import annotations

import json
import logging
import os
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from engines import EngineResult

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# JSON Report
# ---------------------------------------------------------------------------

def write_json_report(merged: dict, job_id: str, output_dir: str) -> str:
    """Write merged output to a JSON file. Returns the file path."""
    reports_dir = os.path.join(output_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    path = os.path.join(reports_dir, f"{job_id}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(merged, f, indent=2, default=str, ensure_ascii=False)
    logger.info("JSON report written: %s", path)
    return path


# ---------------------------------------------------------------------------
# HTML Report (inline Jinja2 template)
# ---------------------------------------------------------------------------

_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Scraper Report — {{ job_id }}</title>
<style>
  :root {
    --bg: #0f1117; --surface: #1a1d27; --surface2: #232636;
    --border: #2e3247; --accent: #6c63ff; --success: #22c55e;
    --warn: #f59e0b; --danger: #ef4444; --text: #e2e8f0;
    --text2: #94a3b8; --code: #1e2233;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { background: var(--bg); color: var(--text); font-family: 'Segoe UI', system-ui, sans-serif; line-height: 1.6; }
  a { color: var(--accent); text-decoration: none; }
  a:hover { text-decoration: underline; }

  .header { background: var(--surface); border-bottom: 1px solid var(--border); padding: 1.5rem 2rem; display: flex; align-items: center; gap: 1rem; }
  .header h1 { font-size: 1.5rem; }
  .header .badge { background: var(--accent); color: #fff; padding: .2rem .7rem; border-radius: 999px; font-size: .75rem; font-weight: 700; }

  .container { max-width: 1200px; margin: 0 auto; padding: 2rem; }

  .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 1.5rem; }
  @media (max-width: 800px) { .grid-2 { grid-template-columns: 1fr; } }

  .card { background: var(--surface); border: 1px solid var(--border); border-radius: .75rem; padding: 1.5rem; margin-bottom: 1.5rem; }
  .card h2 { font-size: 1.1rem; margin-bottom: 1rem; color: var(--accent); display: flex; align-items: center; gap: .5rem; }
  .card h3 { font-size: .95rem; margin-bottom: .5rem; color: var(--text2); }

  .stat-row { display: flex; flex-wrap: wrap; gap: 1rem; margin-bottom: 1.5rem; }
  .stat-box { background: var(--surface2); border: 1px solid var(--border); border-radius: .5rem; padding: .75rem 1.25rem; min-width: 140px; }
  .stat-box .val { font-size: 2rem; font-weight: 800; color: var(--accent); }
  .stat-box .lbl { font-size: .75rem; color: var(--text2); text-transform: uppercase; letter-spacing: .05em; }

  .conf-bar { height: 8px; background: var(--surface2); border-radius: 999px; margin: .25rem 0 .75rem; overflow: hidden; }
  .conf-bar-fill { height: 100%; border-radius: 999px; background: linear-gradient(90deg, var(--accent), var(--success)); }

  table { width: 100%; border-collapse: collapse; font-size: .85rem; }
  th { background: var(--surface2); padding: .6rem .8rem; text-align: left; font-size: .75rem; color: var(--text2); text-transform: uppercase; letter-spacing: .05em; }
  td { padding: .55rem .8rem; border-bottom: 1px solid var(--border); vertical-align: top; word-break: break-word; }
  tr:last-child td { border-bottom: none; }
  .success-cell { color: var(--success); font-weight: 600; }
  .fail-cell { color: var(--danger); font-weight: 600; }

  .tag { display: inline-block; background: var(--surface2); border: 1px solid var(--border); border-radius: 4px; padding: .1rem .45rem; font-size: .75rem; margin: .15rem; }
  .tag.accent { background: rgba(108,99,255,.15); border-color: var(--accent); color: var(--accent); }

  .collapsible summary { cursor: pointer; padding: .5rem 0; font-weight: 600; color: var(--text2); }
  .collapsible summary:hover { color: var(--text); }

  .text-block { background: var(--code); border: 1px solid var(--border); border-radius: .5rem; padding: 1rem; font-family: monospace; font-size: .8rem; white-space: pre-wrap; word-break: break-word; max-height: 300px; overflow-y: auto; color: #a5b4fc; }

  .api-card { background: var(--surface2); border: 1px solid var(--border); border-radius: .5rem; padding: .75rem 1rem; margin-bottom: .5rem; }
  .api-url { font-family: monospace; font-size: .78rem; color: #67e8f9; }

  .link-list { max-height: 250px; overflow-y: auto; }
  .link-list a { display: block; padding: .2rem 0; font-size: .82rem; border-bottom: 1px solid var(--border); color: var(--text2); }
  .link-list a:hover { color: var(--accent); }

  .section-title { font-size: 1.3rem; font-weight: 700; margin: 2rem 0 1rem; border-left: 3px solid var(--accent); padding-left: .75rem; }

  footer { text-align: center; padding: 2rem; color: var(--text2); font-size: .8rem; border-top: 1px solid var(--border); margin-top: 2rem; }

  /* Endpoint probe table */
  .ep-table { width: 100%; border-collapse: collapse; font-size: .8rem; margin-top: .75rem; }
  .ep-table th { background: var(--surface2); padding: .5rem .7rem; text-align: left; font-size: .72rem; color: var(--text2); text-transform: uppercase; letter-spacing: .05em; white-space: nowrap; }
  .ep-table td { padding: .45rem .7rem; border-bottom: 1px solid var(--border); vertical-align: middle; word-break: break-all; }
  .ep-url { font-family: monospace; font-size: .78rem; color: #67e8f9; max-width: 340px; display: block; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .risk-badge { display: inline-block; padding: .1rem .45rem; border-radius: 4px; font-size: .7rem; font-weight: 700; text-transform: uppercase; letter-spacing: .04em; }
  .risk-badge.risk-high   { background: rgba(239,68,68,.18); color: #f87171; border: 1px solid rgba(239,68,68,.4); }
  .risk-badge.risk-medium { background: rgba(251,146,60,.15); color: #fb923c; border: 1px solid rgba(251,146,60,.35); }
  .risk-badge.risk-low    { background: rgba(74,222,128,.12); color: #4ade80; border: 1px solid rgba(74,222,128,.3); }
  .risk-badge.risk-info   { background: rgba(148,163,184,.1); color: #94a3b8; border: 1px solid rgba(148,163,184,.25); }
  .ep-filter-bar { display: flex; gap: .5rem; flex-wrap: wrap; margin-bottom: .75rem; }
  .ep-filter-bar span { display: inline-block; padding: .15rem .5rem; border-radius: 4px; font-size: .75rem; font-weight: 600; cursor: default; }
  .more-hint { color: var(--text2); font-size: .8rem; margin-top: .5rem; }
  .ep-summary-badges { display: flex; flex-wrap: wrap; gap: .5rem; margin-bottom: .85rem; }
  .ep-summary-badge { display: inline-block; padding: .2rem .65rem; border-radius: 999px; font-size: .75rem; font-weight: 600; }
  .ep-badge-green  { background: rgba(34,197,94,.15); color: #4ade80; border: 1px solid rgba(34,197,94,.3); }
  .ep-badge-purple { background: rgba(167,139,250,.15); color: #a78bfa; border: 1px solid rgba(167,139,250,.3); }
  .ep-badge-red    { background: rgba(239,68,68,.15); color: #f87171; border: 1px solid rgba(239,68,68,.3); }
  .ep-badge-gray   { background: rgba(148,163,184,.1); color: #94a3b8; border: 1px solid rgba(148,163,184,.2); }
</style>
</head>
<body>

<div class="header">
  <span style="font-size:2rem">⚡</span>
  <div>
    <h1>Universal Scraper Report</h1>
    <div style="color:var(--text2);font-size:.85rem">{{ url }}</div>
  </div>
  <span class="badge" style="margin-left:auto">{{ job_id }}</span>
</div>

<div class="container">

  <!-- OVERVIEW STATS -->
  <div class="stat-row">
    <div class="stat-box">
      <div class="val">{{ engines_used }}</div>
      <div class="lbl">Engines run</div>
    </div>
    <div class="stat-box">
      <div class="val" style="color:var(--success)">{{ engines_succeeded }}</div>
      <div class="lbl">Succeeded</div>
    </div>
    <div class="stat-box">
      <div class="val">{{ "%.0f"|format(confidence_score * 100) }}%</div>
      <div class="lbl">Confidence</div>
    </div>
    <div class="stat-box">
      <div class="val">{{ total_links }}</div>
      <div class="lbl">Links found</div>
    </div>
    <div class="stat-box">
      <div class="val">{{ total_images }}</div>
      <div class="lbl">Images found</div>
    </div>
    <div class="stat-box">
      <div class="val">{{ api_endpoints }}</div>
      <div class="lbl">API endpoints</div>
    </div>
    <div class="stat-box">
      <div class="val">{{ total_elapsed }}s</div>
      <div class="lbl">Total time</div>
    </div>
  </div>

  <!-- SITE SUMMARY -->
  <div class="card">
    <h2>📋 Site Summary</h2>
    <div class="grid-2">
      <div>
        <h3>Title</h3>
        <p>{{ title or "—" }}</p>
        <h3 style="margin-top:.75rem">Description</h3>
        <p style="color:var(--text2);font-size:.9rem">{{ description or "—" }}</p>
        {% if canonical_url %}
        <h3 style="margin-top:.75rem">Canonical URL</h3>
        <p><a href="{{ canonical_url }}">{{ canonical_url }}</a></p>
        {% endif %}
      </div>
      <div>
        <h3>Site Type (detected)</h3>
        <p><span class="tag accent">{{ site_type }}</span></p>
        {% if language and language != "unknown" %}
        <h3 style="margin-top:.75rem">Language</h3>
        <p>{{ language }}</p>
        {% endif %}
        {% if page_type and page_type != "unknown" %}
        <h3 style="margin-top:.75rem">Page Type</h3>
        <p>{{ page_type }}</p>
        {% endif %}
        {% if keywords %}
        <h3 style="margin-top:.75rem">Keywords</h3>
        <div>{% for kw in keywords[:20] %}<span class="tag">{{ kw }}</span>{% endfor %}</div>
        {% endif %}
      </div>
    </div>
  </div>

  <!-- OVERALL CONFIDENCE -->
  <div class="card">
    <h2>📊 Overall Confidence Score</h2>
    <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.5rem">
      <div style="font-size:2.5rem;font-weight:800;color:var(--accent)">{{ "%.1f"|format(confidence_score * 100) }}%</div>
      <div style="color:var(--text2);font-size:.9rem">Based on {{ engines_succeeded }}/{{ engines_used }} engines succeeding and cross-field agreement.</div>
    </div>
    <div class="conf-bar"><div class="conf-bar-fill" style="width:{{ "%.1f"|format(confidence_score * 100) }}%"></div></div>

    {% if field_confidence %}
    <details class="collapsible" style="margin-top:1rem">
      <summary>Per-field confidence breakdown</summary>
      <table style="margin-top:.75rem">
        <tr><th>Field</th><th>Confidence</th><th></th></tr>
        {% for field, conf in field_confidence.items() %}
        <tr>
          <td>{{ field }}</td>
          <td>{{ "%.1f"|format(conf * 100) }}%</td>
          <td style="width:150px">
            <div class="conf-bar" style="margin:0">
              <div class="conf-bar-fill" style="width:{{ "%.1f"|format(conf * 100) }}%"></div>
            </div>
          </td>
        </tr>
        {% endfor %}
      </table>
    </details>
    {% endif %}
  </div>

  <!-- ENGINE RESULTS TABLE -->
  <div class="section-title">🔧 Engine Results</div>
  <div class="card" style="padding:0;overflow:hidden">
    <table>
      <tr>
        <th>Engine</th><th>Status</th><th>HTTP</th><th>Time</th><th>Notes</th>
      </tr>
      {% for eng in engine_summary %}
      <tr>
        <td>
          <span style="font-family:monospace;font-size:.8rem">{{ eng.engine_id }}</span>
        </td>
        <td>
          {% if eng.success %}
          <span class="success-cell">✓ OK</span>
          {% else %}
          <span class="fail-cell">✗ FAIL</span>
          {% endif %}
        </td>
        <td>{{ eng.status_code or "—" }}</td>
        <td>{{ "%.2f"|format(eng.elapsed_s or 0) }}s</td>
        <td style="color:var(--text2);font-size:.78rem">
          {{ eng.error or (eng.warnings[:1][0] if eng.warnings else "") }}
        </td>
      </tr>
      {% endfor %}
    </table>
  </div>

  <div class="grid-2">

    <!-- HEADINGS -->
    {% if headings %}
    <div class="card">
      <h2>📑 Headings ({{ headings|length }})</h2>
      {% for h in headings[:30] %}
      <div style="padding:.2rem 0;padding-left:{{ (h.level - 1) * 12 }}px;font-size:{{ [1.0, 0.95, 0.9, 0.85, 0.82, 0.8][h.level-1] if h.level <= 6 else 0.8 }}rem">
        <span style="color:var(--accent);font-size:.7rem;margin-right:.35rem">H{{ h.level }}</span>{{ h.text }}
      </div>
      {% endfor %}
      {% if headings|length > 30 %}<div style="color:var(--text2);font-size:.8rem;margin-top:.5rem">+{{ headings|length - 30 }} more…</div>{% endif %}
    </div>
    {% endif %}

    <!-- STRUCTURED DATA -->
    {% if structured_data %}
    <div class="card">
      <h2>🧩 Structured Data</h2>
      {% for dtype, dval in structured_data.items() %}
      <details class="collapsible">
        <summary>{{ dtype }}</summary>
        <div class="text-block">{{ dval | tojson(indent=2) }}</div>
      </details>
      {% endfor %}
    </div>
    {% endif %}

  </div>

  <!-- API PAYLOADS -->
  {% if detected_api_data %}
  <div class="card">
    <h2>📡 Detected API Endpoints ({{ detected_api_data|length }})</h2>
    {% for ep in detected_api_data[:20] %}
    <div class="api-card">
      <div class="api-url">{{ ep.endpoint }}</div>
      {% if ep.payload_summary %}
      <div style="color:var(--text2);font-size:.78rem;margin-top:.25rem">
        Keys: {{ ep.payload_summary if ep.payload_summary is string else ep.payload_summary | join(', ') }}
      </div>
      {% endif %}
    </div>
    {% endfor %}
    {% if detected_api_data|length > 20 %}<div style="color:var(--text2);font-size:.8rem">+{{ detected_api_data|length - 20 }} more…</div>{% endif %}
  </div>
  {% endif %}

  <!-- EXPOSED ENDPOINTS (endpoint_probe engine) -->
  {% if detected_endpoints %}
  <div class="card">
    <h2>🔍 Exposed Endpoints &amp; APIs ({{ detected_endpoints|length }})</h2>

    {% if endpoint_probe_summary %}
    <div class="ep-summary-badges">
      {% if endpoint_probe_summary.openapi_discovered %}
      <span class="ep-summary-badge ep-badge-green">📄 OpenAPI: {{ endpoint_probe_summary.openapi_url or '' }}</span>
      {% endif %}
      {% if endpoint_probe_summary.graphql_discovered %}
      <span class="ep-summary-badge ep-badge-purple">⬡ GraphQL: {{ endpoint_probe_summary.graphql_url or '' }}</span>
      {% endif %}
      {% if endpoint_probe_summary.cors_exposed_count %}
      <span class="ep-summary-badge ep-badge-red">⚠ CORS Wildcard: {{ endpoint_probe_summary.cors_exposed_count }} endpoint{{ 's' if endpoint_probe_summary.cors_exposed_count != 1 else '' }}</span>
      {% endif %}
      {% if endpoint_probe_summary.websocket_endpoints %}
      <span class="ep-summary-badge ep-badge-purple">⚡ WebSocket: {{ endpoint_probe_summary.websocket_endpoints|length }}</span>
      {% endif %}
      {% if endpoint_probe_summary.risk_summary %}
      {% set rs = endpoint_probe_summary.risk_summary %}
      {% if rs.high_count %}<span class="ep-summary-badge ep-badge-red">HIGH: {{ rs.high_count }}</span>{% endif %}
      {% if rs.medium_count %}<span class="ep-summary-badge" style="background:rgba(251,146,60,.15);color:#fb923c;border:1px solid rgba(251,146,60,.3)">MEDIUM: {{ rs.medium_count }}</span>{% endif %}
      {% endif %}
      {% if endpoint_probe_summary.js_files_analyzed %}
      <span class="ep-summary-badge ep-badge-gray">JS files analysed: {{ endpoint_probe_summary.js_files_analyzed }}</span>
      {% endif %}
    </div>
    {% endif %}

    <table class="ep-table">
      <thead>
        <tr>
          <th>URL</th><th>Method</th><th>Source</th>
          <th>Auth</th><th>CORS</th><th>Status</th><th>Risk</th>
        </tr>
      </thead>
      <tbody>
      {% for ep in detected_endpoints[:60] %}
      <tr>
        <td><span class="ep-url" title="{{ ep.url }}">{{ ep.url }}</span></td>
        <td><code style="font-size:.75rem">{{ ep.method }}</code></td>
        <td style="color:var(--text2);font-size:.75rem">{{ ep.source | replace('_', ' ') }}</td>
        <td>
          {% if ep.auth_required == false %}
            <span style="color:#f87171">✗ Open</span>
          {% elif ep.auth_required == true %}
            <span style="color:#4ade80">✓ Auth</span>
          {% else %}
            <span style="color:#94a3b8">?</span>
          {% endif %}
        </td>
        <td>
          {% if ep.cors_permissive %}
            <span style="color:#fb923c">⚠ Wildcard</span>
          {% else %}—{% endif %}
        </td>
        <td>
          {% if ep.status_code %}
            <span style="color:{% if ep.status_code == 200 %}#4ade80{% elif ep.status_code in [401,403] %}#fb923c{% else %}#94a3b8{% endif %}">
              {{ ep.status_code }}
            </span>
          {% else %}—{% endif %}
        </td>
        <td><span class="risk-badge risk-{{ ep.risk_level | lower }}">{{ ep.risk_level }}</span></td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    {% if detected_endpoints|length > 60 %}
    <div class="more-hint">+{{ detected_endpoints|length - 60 }} more endpoints in the JSON report</div>
    {% endif %}
  </div>
  {% endif %}

  <!-- LINKS -->
  {% if links %}
  <div class="card">
    <h2>🔗 Links ({{ links|length }})</h2>
    <div class="link-list">
      {% for l in links[:100] %}
      <a href="{{ l.href }}" target="_blank" rel="noopener noreferrer">
        {{ l.text or l.href }}
      </a>
      {% endfor %}
    </div>
    {% if links|length > 100 %}<div style="color:var(--text2);font-size:.8rem;margin-top:.5rem">+{{ links|length - 100 }} more…</div>{% endif %}
  </div>
  {% endif %}

  <!-- IMAGES -->
  {% if images %}
  <div class="card">
    <h2>🖼️ Images ({{ images|length }})</h2>
    <div style="display:flex;flex-wrap:wrap;gap:.5rem">
      {% for img in images[:30] %}
      <a href="{{ img.src }}" target="_blank" rel="noopener noreferrer">
        <img src="{{ img.src }}" alt="{{ img.alt }}" title="{{ img.title }}"
             style="width:80px;height:60px;object-fit:cover;border-radius:4px;border:1px solid var(--border)"
             onerror="this.style.display='none'">
      </a>
      {% endfor %}
    </div>
    {% if images|length > 30 %}<div style="color:var(--text2);font-size:.8rem;margin-top:.5rem">+{{ images|length - 30 }} more…</div>{% endif %}
  </div>
  {% endif %}

  <!-- MAIN CONTENT PREVIEW -->
  {% if main_content %}
  <div class="card">
    <h2>📄 Main Content Preview</h2>
    <div class="text-block">{{ main_content[:3000] }}{% if main_content|length > 3000 %}
…(truncated){% endif %}</div>
  </div>
  {% endif %}

  <!-- CRAWL MAP -->
  {% if crawl_pages %}
  <div class="card">
    <h2>🕷️ Crawled Pages ({{ crawl_pages|length }})</h2>
    <table>
      <tr><th>URL</th><th>Title</th><th>Depth</th><th>Status</th></tr>
      {% for page in crawl_pages[:50] %}
      <tr>
        <td><a href="{{ page.url }}" target="_blank" style="font-size:.78rem">{{ page.url[:60] }}{% if page.url|length > 60 %}…{% endif %}</a></td>
        <td style="font-size:.82rem">{{ page.title or "—" }}</td>
        <td>{{ page.depth }}</td>
        <td>{{ page.status }}</td>
      </tr>
      {% endfor %}
    </table>
  </div>
  {% endif %}

  <!-- NAVIGATION STRUCTURE -->
  {% if links %}
  <div class="card">
    <h2>🧭 Navigation Structure ({{ links|length }} links)</h2>

    <!-- Tree view: group links by domain -->
    <div id="nav-tree" style="margin-bottom:1rem"></div>
    <script>
    (function(){
      var links = {{ links_json }};
      var byDomain = {};
      links.forEach(function(l){
        var href = l.href || '';
        try {
          var u = new URL(href);
          var d = u.hostname;
          if (!byDomain[d]) byDomain[d] = [];
          byDomain[d].push({path: u.pathname + u.search, text: l.text || href, href: href});
        } catch(e) {}
      });
      var container = document.getElementById('nav-tree');
      var domains = Object.keys(byDomain).sort();
      domains.forEach(function(d){
        var det = document.createElement('details');
        det.className = 'collapsible';
        det.style.marginBottom = '.35rem';
        var sum = document.createElement('summary');
        sum.textContent = d + ' (' + byDomain[d].length + ' links)';
        det.appendChild(sum);
        var ul = document.createElement('ul');
        ul.style.cssText = 'list-style:none;padding-left:1rem;max-height:200px;overflow-y:auto';
        byDomain[d].slice(0, 50).forEach(function(item){
          var li = document.createElement('li');
          li.style.cssText = 'padding:.15rem 0;font-size:.82rem;border-bottom:1px solid var(--border)';
          var a = document.createElement('a');
          a.href = item.href; a.target = '_blank'; a.rel = 'noopener noreferrer';
          a.textContent = item.text || item.path;
          a.style.color = 'var(--text2)';
          li.appendChild(a);
          ul.appendChild(li);
        });
        if (byDomain[d].length > 50) {
          var more = document.createElement('li');
          more.style.cssText = 'color:var(--text2);font-size:.78rem;padding:.3rem 0';
          more.textContent = '+' + (byDomain[d].length - 50) + ' more…';
          ul.appendChild(more);
        }
        det.appendChild(ul);
        container.appendChild(det);
      });
      if (!domains.length) container.style.display = 'none';
    })();
    </script>
  </div>
  {% endif %}

  <!-- DIAGNOSTICS -->
  <div class="section-title">🔍 Diagnostics</div>
  <div class="card">
    <h2>Site Analysis</h2>
    <table>
      <tr><th>Property</th><th>Value</th></tr>
      <tr><td>Site Type</td><td>{{ site_type }}</td></tr>
      <tr><td>Is SPA</td><td>{{ "Yes" if site_analysis.is_spa else "No" }}</td></tr>
      <tr><td>Has API Calls</td><td>{{ "Yes" if site_analysis.has_api_calls else "No" }}</td></tr>
      <tr><td>Has Sitemap</td><td>{{ "Yes" if site_analysis.has_sitemap else "No" }}</td></tr>
      <tr><td>Has RSS Feed</td><td>{{ "Yes" if site_analysis.has_rss_feed else "No" }}</td></tr>
      <tr><td>Initial HTTP Status</td><td>{{ site_analysis.initial_status }}</td></tr>
      <tr><td>Conflicting Fields</td><td>{{ conflicting_fields | join(', ') or "None" }}</td></tr>
      <tr><td>Scraped At</td><td>{{ scraped_at }}</td></tr>
      <tr><td>Total Elapsed</td><td>{{ total_elapsed }}s</td></tr>
    </table>
  </div>

  {% if change_detection %}
  <div class="card" style="border-color:{% if change_detection.changed %}var(--warn){% else %}var(--border){% endif %}">
    <h2 style="color:{% if change_detection.changed %}var(--warn){% else %}var(--success){% endif %}">🔄 Change Detection</h2>
    <table>
      <tr><th>Property</th><th>Value</th></tr>
      <tr><td>Changed Since Last Crawl</td><td>{% if change_detection.changed %}<span style="color:var(--warn)">Yes</span>{% else %}<span style="color:var(--success)">No</span>{% endif %}</td></tr>
      <tr><td>Last Seen</td><td>{{ change_detection.last_seen or "First crawl" }}</td></tr>
      <tr><td>Summary</td><td>{{ change_detection.diff_summary or "—" }}</td></tr>
      {% if change_detection.previous_hash %}
      <tr><td>Previous Hash</td><td style="font-family:monospace;font-size:.78rem">{{ change_detection.previous_hash }}</td></tr>
      {% endif %}
    </table>
  </div>
  {% endif %}

  {% if warnings %}
  <div class="card" style="border-color:var(--warn)">
    <h2 style="color:var(--warn)">⚠️ Warnings ({{ warnings|length }})</h2>
    {% for w in warnings %}
    <div style="color:var(--warn);font-size:.85rem;padding:.2rem 0">• {{ w }}</div>
    {% endfor %}
  </div>
  {% endif %}

</div>

<footer>
  Generated by Universal Multi-Engine Web Scraper &mdash; Job {{ job_id }} &mdash; {{ scraped_at }}
</footer>
</body>
</html>"""


def write_html_report(
    merged: dict,
    engine_results: list[EngineResult],
    job_id: str,
    output_dir: str,
) -> str:
    """Render the HTML report and write to disk. Returns file path."""
    reports_dir = os.path.join(output_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    path = os.path.join(reports_dir, f"{job_id}.html")

    try:
        from jinja2 import Environment
        try:
            from markupsafe import Markup  # preferred: markupsafe is the upstream package
        except ImportError:
            from jinja2 import Markup  # type: ignore[no-redef]  # Jinja2 re-exports it

        def _script_safe_json(obj: Any) -> str:
            """JSON-encode obj and escape </ to prevent </script> injection in inline scripts."""
            raw = json.dumps(obj, default=str, ensure_ascii=False)
            # Replace </ → <\/ so </script> in scraped data can never close the script tag
            return raw.replace("</", "<\\/")

        env = Environment(autoescape=True)
        env.filters["tojson"] = lambda v, indent=None: json.dumps(v, indent=indent, default=str)

        # Gather crawl pages from engine results if available
        crawl_pages: list[dict] = []
        for er in engine_results:
            if er.engine_id == "crawl_discovery" and er.success:
                crawl_pages = (er.data or {}).get("pages", [])
                break

        # Gather navigation links from headless or DOM engines
        nav_links: list[dict] = []
        for er in engine_results:
            if er.engine_id in ("headless_playwright", "dom_interaction") and er.success:
                nav_links = (er.data or {}).get("nav_links", [])
                if nav_links:
                    break

        tmpl = env.from_string(_HTML_TEMPLATE)
        site_analysis = merged.get("site_analysis", {})
        html = tmpl.render(
            job_id=job_id,
            url=merged.get("url", ""),
            title=merged.get("title", ""),
            description=merged.get("description", ""),
            main_content=merged.get("main_content", ""),
            headings=merged.get("headings", []),
            links=merged.get("links", []),
            images=merged.get("images", []),
            structured_data=merged.get("structured_data", {}),
            detected_api_data=merged.get("detected_api_data", []),
            detected_endpoints=merged.get("detected_endpoints", []),
            endpoint_probe_summary=merged.get("endpoint_probe_summary", {}),
            keywords=merged.get("keywords", []),
            canonical_url=merged.get("canonical_url", ""),
            language=merged.get("language", ""),
            page_type=merged.get("page_type", ""),
            confidence_score=merged.get("confidence_score", 0.0),
            field_confidence=merged.get("field_confidence", {}),
            engine_summary=merged.get("engine_summary", []),
            engines_used=merged.get("engines_used", 0),
            engines_succeeded=merged.get("engines_succeeded", 0),
            conflicting_fields=merged.get("conflicting_fields", []),
            warnings=merged.get("warnings", []),
            site_analysis=site_analysis,
            site_type=site_analysis.get("site_type", "unknown"),
            total_links=len(merged.get("links", [])),
            total_images=len(merged.get("images", [])),
            api_endpoints=len(merged.get("detected_api_data", [])),
            scraped_at=merged.get("scraped_at", ""),
            total_elapsed=merged.get("_total_elapsed_s",
                                     sum(e.get("elapsed_s") or 0
                                         for e in merged.get("engine_summary", []))),
            crawl_pages=crawl_pages,
            nav_links=nav_links,
            links_json=Markup(_script_safe_json(merged.get("links", []))),
            change_detection=merged.get("change_detection"),
        )

    except ImportError:
        # Fallback: minimal HTML without Jinja2
        html = f"""<!DOCTYPE html>
<html><head><meta charset='UTF-8'><title>Report {job_id}</title></head>
<body style='font-family:monospace;background:#0f1117;color:#e2e8f0;padding:2rem'>
<h1>⚡ Scraper Report — {job_id}</h1>
<h2>URL: {merged.get('url','')}</h2>
<pre>{json.dumps(merged, indent=2, default=str)[:50000]}</pre>
</body></html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    logger.info("HTML report written: %s", path)
    return path


# ---------------------------------------------------------------------------
# CSV Report
# ---------------------------------------------------------------------------

def write_csv_report(merged: dict, job_id: str, output_dir: str) -> str:
    """
    Write a flat CSV of all discovered links.
    Returns file path, or empty string on error.
    """
    try:
        import csv
        reports_dir = os.path.join(output_dir, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        path = os.path.join(reports_dir, f"{job_id}.csv")

        links = merged.get("links", [])
        images = merged.get("images", [])
        headings = merged.get("headings", [])

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["type", "src_or_href", "text_or_alt", "extra"])
            for lnk in links:
                if isinstance(lnk, dict):
                    writer.writerow(["link", lnk.get("href", ""), lnk.get("text", ""), ""])
                else:
                    writer.writerow(["link", str(lnk), "", ""])
            for img in images:
                if isinstance(img, dict):
                    writer.writerow(["image", img.get("src", ""), img.get("alt", ""),
                                     img.get("title", "")])
                else:
                    writer.writerow(["image", str(img), "", ""])
            for h in headings:
                if isinstance(h, dict):
                    writer.writerow(["heading", "", h.get("text", ""),
                                     f"H{h.get('level', '?')}"])
                else:
                    writer.writerow(["heading", "", str(h), ""])

        logger.info("CSV report written: %s", path)
        return path
    except Exception as exc:
        logger.warning("CSV report generation failed: %s", exc)
        return ""


# ---------------------------------------------------------------------------
# XLSX Report
# ---------------------------------------------------------------------------

def write_xlsx_report(merged: dict, job_id: str, output_dir: str) -> str:
    """
    Write a multi-sheet XLSX report (links, images, headings, metadata).
    Requires openpyxl. Returns file path or empty string on error.
    """
    try:
        import openpyxl  # noqa: F401  (import test)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        reports_dir = os.path.join(output_dir, "reports")
        os.makedirs(reports_dir, exist_ok=True)
        path = os.path.join(reports_dir, f"{job_id}.xlsx")

        wb = Workbook()
        _HEADER_FONT = Font(bold=True, color="FFFFFF")
        _HEADER_FILL = PatternFill("solid", fgColor="4C4F8A")

        def _header_row(ws, cols: list[str]) -> None:
            ws.append(cols)
            for cell in ws[1]:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL

        # Sheet 1 — Links
        ws_links = wb.active
        ws_links.title = "Links"
        _header_row(ws_links, ["href", "text", "rel", "is_external"])
        for lnk in (merged.get("links") or []):
            if isinstance(lnk, dict):
                ws_links.append([
                    lnk.get("href", ""),
                    lnk.get("text", ""),
                    lnk.get("rel", ""),
                    lnk.get("is_external", ""),
                ])
            else:
                ws_links.append([str(lnk), "", "", ""])

        # Sheet 2 — Images
        ws_img = wb.create_sheet("Images")
        _header_row(ws_img, ["src", "alt", "title", "width", "height"])
        for img in (merged.get("images") or []):
            if isinstance(img, dict):
                ws_img.append([
                    img.get("src", ""),
                    img.get("alt", ""),
                    img.get("title", ""),
                    img.get("width", ""),
                    img.get("height", ""),
                ])
            else:
                ws_img.append([str(img), "", "", "", ""])

        # Sheet 3 — Headings
        ws_h = wb.create_sheet("Headings")
        _header_row(ws_h, ["level", "text", "id"])
        for h in (merged.get("headings") or []):
            if isinstance(h, dict):
                ws_h.append([h.get("level", ""), h.get("text", ""), h.get("id", "")])
            else:
                ws_h.append(["", str(h), ""])

        # Sheet 4 — Metadata
        ws_meta = wb.create_sheet("Metadata")
        _header_row(ws_meta, ["key", "value"])
        meta_fields = [
            "title", "description", "canonical_url", "language", "page_type",
            "content_hash", "confidence_score", "engines_used", "engines_succeeded",
            "scraped_at",
        ]
        for k in meta_fields:
            v = merged.get(k)
            if v is not None:
                ws_meta.append([k, str(v)])

        wb.save(path)
        logger.info("XLSX report written: %s", path)
        return path
    except ImportError:
        logger.warning("openpyxl not installed — XLSX export unavailable")
        return ""
    except Exception as exc:
        logger.warning("XLSX report generation failed: %s", exc)
        return ""


# ---------------------------------------------------------------------------
# GraphML crawl map
# ---------------------------------------------------------------------------

def write_crawl_graph(merged: dict, job_id: str, output_dir: str) -> str:
    """
    Write a GraphML file representing the crawl link graph.
    Requires networkx.  Falls back to a minimal hand-written GraphML if
    networkx is missing.
    Returns file path or empty string on error.
    """
    reports_dir = os.path.join(output_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    path = os.path.join(reports_dir, f"{job_id}.graphml")

    # Build edge list: root URL → each discovered internal/external link
    root_url = merged.get("url", "")
    links = merged.get("links", []) or []
    internal_links: list[str] = merged.get("site_analysis", {}).get("internal_links", [])

    edges: list[tuple[str, str]] = []
    for lnk in links:
        href = lnk.get("href", "") if isinstance(lnk, dict) else str(lnk)
        if href and href.startswith("http"):
            edges.append((root_url, href))
    for href in internal_links:
        if href and href.startswith("http"):
            edges.append((root_url, href))

    try:
        import networkx as nx
        G = nx.DiGraph()
        G.add_node(root_url, label="root", job_id=job_id)
        for src, dst in edges:
            G.add_node(dst)
            G.add_edge(src, dst)
        nx.write_graphml(G, path)
        logger.info("GraphML crawl map written: %s (%d nodes, %d edges)",
                    path, G.number_of_nodes(), G.number_of_edges())
        return path

    except ImportError:
        # Minimal hand-rolled GraphML (no networkx needed)
        from xml.sax.saxutils import escape as _esc
        lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<graphml xmlns="http://graphml.graphdrawing.org/graphml">',
            '<graph id="crawl" edgedefault="directed">',
            f'  <node id="{_esc(root_url)}"/>',
        ]
        seen_nodes = {root_url}
        for i, (src, dst) in enumerate(edges):
            if dst not in seen_nodes:
                lines.append(f'  <node id="{_esc(dst)}"/>')
                seen_nodes.add(dst)
            lines.append(f'  <edge id="e{i}" source="{_esc(src)}" target="{_esc(dst)}"/>')
        lines += ['</graph>', '</graphml>']
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        logger.info("GraphML (no-networkx) written: %s", path)
        return path

    except Exception as exc:
        logger.warning("GraphML generation failed: %s", exc)
        return ""


# ---------------------------------------------------------------------------
# Write all reports in one call
# ---------------------------------------------------------------------------

def write_all_reports(
    merged: dict,
    engine_results: Any,
    job_id: str,
    output_dir: str,
) -> dict[str, str]:
    """Convenience wrapper — write JSON, HTML, CSV, XLSX, and GraphML."""
    paths: dict[str, str] = {}
    paths["json"] = write_json_report(merged, job_id, output_dir)
    paths["html"] = write_html_report(merged, engine_results, job_id, output_dir)
    paths["csv"] = write_csv_report(merged, job_id, output_dir)
    paths["xlsx"] = write_xlsx_report(merged, job_id, output_dir)
    paths["graphml"] = write_crawl_graph(merged, job_id, output_dir)
    return paths
